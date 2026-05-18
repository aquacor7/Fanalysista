"""
Download formations xlsx for every giornata in a (league, competition).

Resumes by default: if a giornata's xlsx already exists in the bronze folder
and is non-trivial size, it's skipped. Pass --force to re-download everything.

If the user is the league admin, the rounds list comes from the /Giornate
endpoint. Otherwise (you're a regular member of the league) /Giornate is
permission-blocked, so the script falls back to probing the download endpoint
itself — try giornata 1, 2, 3... until three in a row come back empty.

Usage:
    python download_all.py -l "My League" -c "Serie C"
    python download_all.py -l "My League" -c "Serie C" -t "My Team"
    python download_all.py -l "My League" -c "Serie C" --force
"""
import argparse
import time
from pathlib import Path

from openpyxl import load_workbook

from cli import add_common_args, login_and_select

# When /Giornate isn't available, we probe this many rounds before giving up,
# stopping early after consecutive empties.
DEFAULT_MAX_ROUND = 40
EMPTY_RUN_STOP = 3  # consecutive empties => assume past the end of the season


def _is_placeholder(path: Path) -> bool:
    """Return True if the xlsx is the server's 'file non disponibile' stub.

    For rounds that don't exist the server still returns HTTP 200 with a small
    (~2.6 KB) xlsx whose only cell is ``A1 = "File non disponibile."`` and
    whose sheet is named ``"Leghe Fantacalcio"``. Real giornate xlsx have a
    sheet name like ``"Formazioni N giornata"`` and proper formation data.
    Byte-size alone isn't a clean signal — check the content.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return False  # if it doesn't parse, leave it for downstream review
    a1 = ws["A1"].value
    if isinstance(a1, str) and "non disponibile" in a1.lower():
        return True
    if not (ws.title or "").lower().startswith("formazioni"):
        return True
    return False


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    add_common_args(parser)
    parser.add_argument(
        "--min-bytes", type=int, default=1024,
        help="Skip downloads smaller than this many bytes (default: 1024)",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Re-download even if a file for that giornata already exists",
    )
    parser.add_argument(
        "--max-round", type=int, default=DEFAULT_MAX_ROUND,
        help=(f"Upper bound when probing rounds (default: {DEFAULT_MAX_ROUND}). "
              "Only used when /Giornate is unavailable (e.g. you're not "
              "this league's admin)."),
    )
    args = parser.parse_args()

    client, _league, comp, out_dir = login_and_select(args)

    rounds = client.get_competition_rounds(comp.id)
    if rounds:
        probe_mode = False
        print(f"[rounds]     {len(rounds)} configured: {rounds[0]}..{rounds[-1]}")
    else:
        probe_mode = True
        print(f"[rounds]     /Giornate returned nothing — probably because "
              f"you're not this league's admin. Probing 1..{args.max_round} "
              f"via the download endpoint; stops after {EMPTY_RUN_STOP} "
              f"consecutive empties.")
        rounds = list(range(1, args.max_round + 1))

    out_dir.mkdir(parents=True, exist_ok=True)

    saved, skipped_existing, skipped_empty, failed = [], [], [], []
    consecutive_empty = 0

    for r in rounds:
        # The server's filename pattern is "Formazioni_{alias}_{N}_giornata.xlsx".
        # Glob locally to detect a previous successful download.
        existing = list(out_dir.glob(f"Formazioni_*_{r}_giornata.xlsx"))
        if existing and not args.force:
            f = existing[0]
            if _is_placeholder(f):
                # A previous run kept a placeholder; clean it up and count
                # it as a confirmed empty round.
                print(f"  g{r:>2}: stale placeholder on disk ({f.name}) — removing.")
                f.unlink()
                skipped_empty.append(r)
                consecutive_empty += 1
                if probe_mode and consecutive_empty >= EMPTY_RUN_STOP:
                    print(f"  ... {EMPTY_RUN_STOP} consecutive empties — stopping early.")
                    break
                continue
            if f.stat().st_size >= args.min_bytes:
                print(f"  g{r:>2}: already on disk ({f.name}, "
                      f"{f.stat().st_size} B) — skipping. Use --force to re-download.")
                skipped_existing.append(r)
                consecutive_empty = 0
                continue

        t0 = time.monotonic()
        print(f"  g{r:>2}: downloading...", end=" ")
        try:
            path = client.download_formations(comp.id, r, out_dir=out_dir)
        except Exception as e:
            print(f"FAILED - {type(e).__name__}: {e}")
            failed.append(r)
            continue

        size = path.stat().st_size
        elapsed = time.monotonic() - t0
        empty = size < args.min_bytes or _is_placeholder(path)
        if empty:
            path.unlink()
            print(f"placeholder ({size} B in {elapsed:.1f}s), discarded")
            skipped_empty.append(r)
            consecutive_empty += 1
            if probe_mode and consecutive_empty >= EMPTY_RUN_STOP:
                print(f"  ... {EMPTY_RUN_STOP} consecutive empties — "
                      f"assuming end of season. Stopping early.")
                break
        else:
            print(f"{path.name} ({size} B in {elapsed:.1f}s)")
            saved.append(r)
            consecutive_empty = 0

    print()
    print(f"[done] saved={len(saved)} "
          f"skipped_existing={len(skipped_existing)} "
          f"skipped_empty={len(skipped_empty)} "
          f"failed={len(failed)}")
    print(f"       out: {out_dir}/")
    if failed:
        print(f"       retry failed giornate: {failed}")


if __name__ == "__main__":
    main()
