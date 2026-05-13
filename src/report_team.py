"""
Build a per-player wide-format xlsx report for one team across all giornata.

Reads silver/{league}/{comp}/appearances.csv (built by build_silver.py)
and writes a presentation-friendly xlsx to:
    reports/{league}/{comp}/{team}.xlsx

Layout:
    rows  = unique players who appeared for the team
    cols  = position | player | (voto, fantavoto, active) x N giornate
            | apps_active | total_active_fv | avg_active_fv
            | total_fv_missed | avg_fv_missed | total_voto | total_fv

Inactive (greyed-out) appearances are shaded; players who didn't appear at all
in a giornata leave blank cells.

Usage:
    python report_team.py -l My League -c "Serie C" -t "My Team"
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cli import resolve_layer_dir, safe_slug

POSITION_ORDER = {"P": 0, "D": 1, "C": 2, "A": 3}

INACTIVE_FILL = PatternFill("solid", fgColor="FFEFEFEF")
INACTIVE_FONT = Font(color="FF888888", italic=True)
HEADER_FILL = PatternFill("solid", fgColor="FFE8E8E8")
HEADER_FONT = Font(bold=True)
SUMMARY_FONT = Font(bold=True)


def build_report(appearances: pd.DataFrame, matches: pd.DataFrame, team: str, out_path: Path) -> None:
    team_ap = appearances[appearances.team.str.lower() == team.lower()].copy()
    if team_ap.empty:
        raise LookupError(f"team {team!r} not found in appearances")
    team_mt = matches[matches.team.str.lower() == team.lower()].copy()

    giornate = sorted(team_ap.giornata.unique().tolist())

    # Canonical position per player (first-seen)
    canon = (team_ap.sort_values("giornata")
                    .drop_duplicates("player", keep="first")
                    .set_index("player")["position"]
                    .to_dict())

    # Sort by position then by name
    players = sorted(canon.keys(), key=lambda n: (POSITION_ORDER.get(canon[n], 9), n))

    # Index: (player, giornata) -> row
    ap_idx = team_ap.set_index(["player", "giornata"])

    wb = Workbook()
    ws = wb.active
    ws.title = safe_slug(team)[:31]

    # ---- header ----
    ws.cell(1, 1, "Pos").font = HEADER_FONT
    ws.cell(1, 2, "Player").font = HEADER_FONT
    for i, g in enumerate(giornate):
        base = 3 + i * 3
        for j, label in enumerate(("voto", "fv", "active")):
            c = ws.cell(1, base + j, f"g{g}_{label}")
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
    tail = 3 + len(giornate) * 3
    summary_labels = (
        "apps_active", "total_active_fv", "avg_active_fv",
        "total_fv_missed", "avg_fv_missed",
        "total_voto", "total_fv",
    )
    for i, label in enumerate(summary_labels):
        ws.cell(1, tail + i, label).font = HEADER_FONT

    # ---- player rows ----
    for ri, name in enumerate(players, start=2):
        ws.cell(ri, 1, canon[name])
        ws.cell(ri, 2, name)

        apps_active = 0
        total_active_fv = 0.0
        missed_apps = 0
        total_fv_missed = 0.0
        total_voto = 0.0
        total_fv = 0.0

        for i, g in enumerate(giornate):
            try:
                row = ap_idx.loc[(name, g)]
            except KeyError:
                continue
            voto = None if pd.isna(row.voto) else float(row.voto)
            fv = None if pd.isna(row.fantavoto) else float(row.fantavoto)
            active = bool(row.active)

            base = 3 + i * 3
            v_cell = ws.cell(ri, base + 0, voto)
            fv_cell = ws.cell(ri, base + 1, fv)
            a_cell = ws.cell(ri, base + 2, "Y" if active else "N")
            if not active:
                for c in (v_cell, fv_cell, a_cell):
                    c.fill = INACTIVE_FILL
                    c.font = INACTIVE_FONT

            if voto is not None:
                total_voto += voto
            if fv is not None:
                total_fv += fv
            if active:
                apps_active += 1
                if fv is not None:
                    total_active_fv += fv
            elif fv is not None:
                missed_apps += 1
                total_fv_missed += fv

        values = (
            apps_active,
            round(total_active_fv, 2),
            round(total_active_fv / apps_active, 2) if apps_active else None,
            round(total_fv_missed, 2),
            round(total_fv_missed / missed_apps, 2) if missed_apps else None,
            round(total_voto, 2),
            round(total_fv, 2),
        )
        for i, val in enumerate(values):
            ws.cell(ri, tail + i, val).font = SUMMARY_FONT

    # ---- TOTALE row at the bottom (from matches.csv) ----
    bottom = len(players) + 2
    ws.cell(bottom, 2, "TOTALE").font = HEADER_FONT
    tot_by_g = team_mt.set_index("giornata")["totale"].to_dict()
    for i, g in enumerate(giornate):
        tot = tot_by_g.get(g)
        if tot is not None:
            c = ws.cell(bottom, 3 + i * 3 + 1, float(tot))
            c.font = HEADER_FONT
            c.fill = HEADER_FILL

    # ---- column widths / freeze ----
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for i in range(len(giornate)):
        for j in range(3):
            ws.column_dimensions[get_column_letter(3 + i * 3 + j)].width = 6 if j < 2 else 7
    for i in range(len(summary_labels)):
        ws.column_dimensions[get_column_letter(tail + i)].width = 14
    ws.freeze_panes = "C2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-l", "--league", required=True)
    parser.add_argument("-c", "--competition", required=True)
    parser.add_argument("-t", "--team", required=True)
    parser.add_argument("--silver", default="silver")
    parser.add_argument("--reports", default="reports")
    args = parser.parse_args()

    silver_dir, league_alias, comp_slug = resolve_layer_dir(
        Path(args.silver), args.league, args.competition,
    )
    print(f"[silver] {silver_dir}/")

    appearances = pd.read_csv(silver_dir / "appearances.csv")
    matches = pd.read_csv(silver_dir / "matches.csv")
    print(f"         appearances={len(appearances)} matches={len(matches)}")

    out_path = Path(args.reports) / league_alias / comp_slug / f"{safe_slug(args.team)}.xlsx"
    build_report(appearances, matches, args.team, out_path)
    print(f"[report] {out_path}  ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
