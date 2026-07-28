"""
Parse the two auction/market bronze files into tidy records.

1. Quotazioni_Fantacalcio_Stagione_YYYY_NN.xlsx — one row per real Serie A
   footballer, with the season's price quotations. Layout (sheet "Tutti"):
       row 1: title
       row 2: header  [Id, R, RM, Nome, Squadra, Qt.A, Qt.I, Diff., Qt.A M,
                        Qt.I M, Diff.M, FVM, FVM M]
       row 3+: data
   We keep the Classic-mode columns (R, Nome, Squadra, Qt.A, Qt.I, Diff., FVM)
   and drop the four Mantra-mode columns (…M).

     - Qt.A  = current quotation (as of the last giornata)
     - Qt.I  = initial quotation (first giornata)
     - Diff. = Qt.A − Qt.I  (season value change)
     - FVM   = FantaValore di Mercato — a market-consensus "fair" auction price
               on a 1000-credit budget.

2. Rose_{league}.xlsx — each participant's purchased squad (auction result).
   The sheet stacks participant blocks two-across:
       [participant name] .......... [participant name]     (cols A / F)
       Ruolo Calciatore Squadra Costo | Ruolo Calciatore Squadra Costo
       <player rows: role in {P,D,C,A}>
       ...
       Crediti Residui: N .......... Crediti Residui: N
   Left block lives in cols A–D, right block in cols F–I, col E is a gutter.

Player names carry the same trailing-"*" (left-the-pool) marker as the
formation files; we normalise it away so joins line up.
"""
from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

_TRAILING_STAR = re.compile(r"\s*\*+\s*$")
_VALID_ROLES = {"P", "D", "C", "A"}
_CREDITS_RE = re.compile(r"Crediti\s+Residui:\s*(-?\d+)", re.IGNORECASE)


def _canonical_name(raw) -> str:
    return _TRAILING_STAR.sub("", str(raw or "")).strip()


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


# ---------------- Quotazioni ----------------

def parse_quotazioni(path: Path) -> pd.DataFrame:
    """Return one row per footballer: player, role, club, qt_a, qt_i, val_diff, fvm."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Tutti"] if "Tutti" in wb.sheetnames else wb.active

    # Find the header row (the one whose first cell is "Id").
    header_row = next(
        (r for r in range(1, 6) if str(ws.cell(row=r, column=1).value).strip() == "Id"),
        2,
    )
    header = [str(ws.cell(row=header_row, column=c).value).strip()
              for c in range(1, ws.max_column + 1)]
    idx = {name: i + 1 for i, name in enumerate(header)}

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=idx["Nome"]).value
        if not name:
            continue
        rows.append({
            "player": _canonical_name(name),
            "role": ws.cell(row=r, column=idx["R"]).value,
            "club": ws.cell(row=r, column=idx["Squadra"]).value,
            "qt_a": _num(ws.cell(row=r, column=idx["Qt.A"]).value),
            "qt_i": _num(ws.cell(row=r, column=idx["Qt.I"]).value),
            "val_diff": _num(ws.cell(row=r, column=idx["Diff."]).value),
            "fvm": _num(ws.cell(row=r, column=idx["FVM"]).value),
        })
    df = pd.DataFrame(rows)
    # A player can appear once per role only; keep the first (Tutti has no dups
    # but Ceduti players may repeat — de-dupe defensively on (player, role)).
    return df.drop_duplicates(subset=["player", "role"]).reset_index(drop=True)


# ---------------- Rose (rosters) ----------------

def parse_rosters(path: Path) -> pd.DataFrame:
    """Return one row per (participant, player): participant, role, player, club, costo.

    Also carries `credits_residui` — the participant's leftover budget — repeated
    on every one of their rows (so a groupby can recover budget = spent + residui).
    """
    ws = load_workbook(path, data_only=True).active
    grid = [[ws.cell(row=r, column=c).value for c in range(1, 10)]
            for r in range(1, ws.max_row + 1)]
    n = len(grid)

    out: list[dict] = []
    residui: dict[str, float | None] = {}
    left = right = None  # active participant names for the two columns

    i = 0
    while i < n:
        row = grid[i]
        # A name row is immediately followed by a "Ruolo" header row.
        if i + 1 < n and grid[i + 1][0] == "Ruolo":
            left = str(row[0]).strip() if row[0] else None
            right = str(row[5]).strip() if row[5] else None
            i += 2
            continue

        for name, (role_c, player_c, club_c, costo_c) in (
            (left, (row[0], row[1], row[2], row[3])),
            (right, (row[5], row[6], row[7], row[8])),
        ):
            if name is None:
                continue
            if role_c in _VALID_ROLES:
                out.append({
                    "participant": name,
                    "role": role_c,
                    "player": _canonical_name(player_c),
                    "club": club_c,
                    "costo": _num(costo_c),
                })
            elif isinstance(role_c, str):
                m = _CREDITS_RE.search(role_c)
                if m:
                    residui[name] = float(m.group(1))
        i += 1

    df = pd.DataFrame(out)
    if df.empty:
        return df
    df["credits_residui"] = df["participant"].map(residui)
    return df
