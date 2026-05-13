"""
Parse a downloaded Formazioni_{league}_{N}_giornata.xlsx into structured records.

xlsx structure: one sheet, matches stacked vertically. Each match has two teams
side-by-side:
  - left team in cols A,B,D,E    (position, name, voto, fantavoto)
  - right team in cols G,H,J,K   (same)
  - score in col F on the match-header row, e.g. "2-0"

Per-team block layout:
    TEAM NAME
    module ("343", "442", ...)
    11 starting player rows (pos in {P,D,C,A})
    'Panchina' row
    N bench rows
    optional 'Modificatore difesa' / 'Fattore campo' rows
    'TOTALE: NN,NN' row
    'Inserita via web il ...' / 'Recuperata dal sistema ...' row

Inactive (didn't count toward TOTALE) players have name font color FFD3D3D3.
A trailing "*" on a player name means "left the player pool" — normalised away
since "Carboni V." and "Carboni V. *" are the same person.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

GREY_FONT = "FFD3D3D3"
VALID_POSITIONS = {"P", "D", "C", "A"}
SCORE_RE = re.compile(r"^\s*(\d+)\s*-\s*(\d+)\s*$")
TOTALE_RE = re.compile(r"TOTALE:\s*([\d,\.]+)")
GIORNATA_RE = re.compile(r"_(\d+)_giornata", re.IGNORECASE)
TRAILING_STAR_RE = re.compile(r"\s*\*+\s*$")


@dataclass
class PlayerRow:
    position: str
    name: str
    voto: Optional[float]
    fantavoto: Optional[float]
    active: bool
    on_bench: bool


@dataclass
class TeamFormation:
    team: str
    side: str                   # 'left' or 'right'
    module: str
    totale: Optional[float]
    modificatore_difesa: Optional[float]
    fattore_campo: Optional[float]
    players: list[PlayerRow] = field(default_factory=list)


@dataclass
class Match:
    giornata: int
    score_left: Optional[int]
    score_right: Optional[int]
    left: TeamFormation
    right: TeamFormation


# -------- low-level helpers --------

def _is_grey_font(cell) -> bool:
    f = cell.font
    return bool(f and f.color and f.color.rgb == GREY_FONT)


def _to_number(v) -> Optional[float]:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _canonical_name(raw: str) -> str:
    return TRAILING_STAR_RE.sub("", raw or "").strip()


# Column layout per side
_SIDE_COLS = {
    "left":  ("A", "B", "D", "E"),
    "right": ("G", "H", "J", "K"),
}


def _parse_team_block(ws: Worksheet, hdr_row: int, side: str) -> TeamFormation:
    """Parse one team's block. Reads from hdr_row downward until TOTALE."""
    pos_c, name_c, voto_c, fv_c = _SIDE_COLS[side]

    team_name = str(ws[f"{pos_c}{hdr_row}"].value or "").strip()
    module = str(ws[f"{pos_c}{hdr_row + 1}"].value or "")

    tf = TeamFormation(
        team=team_name,
        side=side,
        module=module,
        totale=None,
        modificatore_difesa=None,
        fattore_campo=None,
    )

    on_bench = False
    for r in range(hdr_row + 2, ws.max_row + 1):
        pos_val = ws[f"{pos_c}{r}"].value
        if isinstance(pos_val, str):
            stripped = pos_val.strip()
            if stripped == "Panchina":
                on_bench = True
                continue
            if stripped.startswith("TOTALE"):
                m = TOTALE_RE.search(stripped)
                if m:
                    tf.totale = float(m.group(1).replace(",", "."))
                break
            if stripped == "Modificatore difesa":
                # the value lives in the voto column (E for left, K for right)
                tf.modificatore_difesa = _to_number(ws[f"{fv_c}{r}"].value)
                continue
            if stripped == "Fattore campo":
                tf.fattore_campo = _to_number(ws[f"{fv_c}{r}"].value)
                continue

        if pos_val not in VALID_POSITIONS:
            continue

        name_cell = ws[f"{name_c}{r}"]
        tf.players.append(PlayerRow(
            position=pos_val,
            name=_canonical_name(str(name_cell.value or "")),
            voto=_to_number(ws[f"{voto_c}{r}"].value),
            fantavoto=_to_number(ws[f"{fv_c}{r}"].value),
            active=not _is_grey_font(name_cell),
            on_bench=on_bench,
        ))
    return tf


def _find_match_headers(ws: Worksheet) -> list[int]:
    """Return row numbers where a match header lives — col F holds an 'X-Y' score
    AND cols A and G hold non-empty strings."""
    rows = []
    for r in range(1, ws.max_row + 1):
        score = ws[f"F{r}"].value
        if not isinstance(score, str) or not SCORE_RE.match(score):
            continue
        if not (ws[f"A{r}"].value and ws[f"G{r}"].value):
            continue
        rows.append(r)
    return rows


# -------- public API --------

def parse_all_matches(path: Path) -> list[Match]:
    """Return every match in the workbook, both teams parsed."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    m = GIORNATA_RE.search(path.name)
    giornata = int(m.group(1)) if m else 0

    matches: list[Match] = []
    for hdr in _find_match_headers(ws):
        score = SCORE_RE.match(str(ws[f"F{hdr}"].value))
        left = _parse_team_block(ws, hdr, "left")
        right = _parse_team_block(ws, hdr, "right")
        matches.append(Match(
            giornata=giornata,
            score_left=int(score.group(1)) if score else None,
            score_right=int(score.group(2)) if score else None,
            left=left,
            right=right,
        ))
    return matches


def parse_team_in_file(path: Path, team_name: str) -> Optional[TeamFormation]:
    """Convenience wrapper: find one team in one xlsx."""
    target = team_name.strip().lower()
    for match in parse_all_matches(path):
        for tf in (match.left, match.right):
            if tf.team.strip().lower() == target:
                return tf
    return None
